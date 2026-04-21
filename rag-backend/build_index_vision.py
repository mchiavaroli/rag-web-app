# build_index_vision.py
"""
RAG Multimodale + Contextual Retrieval con Azure AI Vision
Combina:
- Contextual Retrieval per chunk di testo (Anthropic 2024)
- Analisi immagini con Azure AI Vision (Image Analysis 4.0)
- Contestualizzazione delle descrizioni immagini
"""
import os, glob, json, re, argparse, base64
from io import BytesIO
from tqdm import tqdm
from pypdf import PdfReader
from PIL import Image
import fitz  # PyMuPDF per estrazione immagini
from sentence_transformers import SentenceTransformer
import numpy as np
import faiss
from config import (MODEL_PROMPT, AZURE_VISION, CHUNK_SIZE, CHUNK_OVERLAP,
                    BATCH_SIZE, MIN_IMAGE_SIZE, USE_LAYOUT_DETECTION, EMBEDDING_MODEL,
                    get_index_path, get_metadata_path, get_chunks_path, get_images_folder,
                    ensure_output_dir)
from llm_client import call_llm_text

# Azure AI Vision
from azure.ai.vision.imageanalysis import ImageAnalysisClient
from azure.ai.vision.imageanalysis.models import VisualFeatures
from azure.core.credentials import AzureKeyCredential


# ===== AZURE AI VISION CLIENT =====

def get_vision_client():
    """Crea il client Azure AI Vision"""
    return ImageAnalysisClient(
        endpoint=AZURE_VISION['endpoint'],
        credential=AzureKeyCredential(AZURE_VISION['api_key'])
    )


def analyze_image_with_azure_vision(image_path, page_num, source_doc, page_text=None):
    """
    Analizza immagine con Azure AI Vision
    Combina: caption, dense captions, tags e OCR
    """
    client = get_vision_client()
    
    try:
        # Leggi l'immagine
        with open(image_path, "rb") as f:
            image_data = f.read()
        
        # Configura le feature da estrarre
        visual_features = [
            VisualFeatures.CAPTION,
            VisualFeatures.DENSE_CAPTIONS,
            VisualFeatures.TAGS,
            VisualFeatures.READ  # OCR
        ]
        
        # Analizza l'immagine
        # Nota: Caption e Dense Captions supportano solo 'en', quindi non passiamo language
        result = client.analyze(
            image_data=image_data,
            visual_features=visual_features,
            gender_neutral_caption=AZURE_VISION.get('gender_neutral_caption', True)
        )
        
        # Costruisci la descrizione combinando tutti i risultati
        description_parts = []
        
        # 1. Caption principale
        if result.caption:
            caption_text = result.caption.text
            confidence = result.caption.confidence
            description_parts.append(f"📷 **Descrizione principale** (confidenza {confidence:.0%}): {caption_text}")
        
        # 2. Dense Captions (descrizioni dettagliate di regioni)
        if result.dense_captions and result.dense_captions.list:
            dense_texts = []
            for dc in result.dense_captions.list[:5]:  # Max 5 dense captions
                if dc.confidence >= 0.5:  # Solo se confidenza >= 50%
                    dense_texts.append(f"- {dc.text} (confidenza {dc.confidence:.0%})")
            if dense_texts:
                description_parts.append(f"\n📍 **Dettagli regioni**:\n" + "\n".join(dense_texts))
        
        # 3. Tags (etichette)
        if result.tags and result.tags.list:
            high_conf_tags = [t.name for t in result.tags.list if t.confidence >= 0.6]
            if high_conf_tags:
                description_parts.append(f"\n🏷️ **Tag**: {', '.join(high_conf_tags[:10])}")
        
        # 4. OCR - Testo estratto dall'immagine
        if result.read and result.read.blocks:
            ocr_texts = []
            for block in result.read.blocks:
                for line in block.lines:
                    text = line.text.strip()
                    if text and len(text) > 2:
                        ocr_texts.append(text)
            if ocr_texts:
                ocr_combined = " | ".join(ocr_texts[:20])  # Max 20 righe
                description_parts.append(f"\n📝 **Testo nell'immagine**: {ocr_combined}")
        
        # 5. Aggiungi contesto pagina
        description_parts.append(f"\n📄 **Posizione**: Pagina {page_num} del documento '{source_doc}'")
        
        # 6. Se disponibile, aggiungi riferimento al contesto della pagina
        if page_text and page_text.strip():
            # Estrai keywords dal testo della pagina
            page_words = set(word.lower() for word in re.findall(r'\b\w{4,}\b', page_text[:2000]))
            # Cerca corrispondenze con i tags
            if result.tags and result.tags.list:
                tag_words = set(t.name.lower() for t in result.tags.list)
                matching = page_words & tag_words
                if matching:
                    description_parts.append(f"\n🔗 **Correlazioni con la pagina**: {', '.join(list(matching)[:5])}")
        
        description = "\n".join(description_parts)
        
        # Log per debug
        print(f"      ✓ Analizzata: {os.path.basename(image_path)}")
        
        return description
        
    except Exception as e:
        print(f"      [WARN] Errore Azure Vision per {image_path}: {e}")
        return f"[Immagine dalla pagina {page_num}] Analisi Azure Vision non disponibile. Errore: {str(e)}"


# ===== FUNZIONI ESTRAZIONE IMMAGINI =====

def extract_images_from_pdf(pdf_path, output_folder=None, min_size=100, use_layout_detection=True):
    """
    Estrae immagini da PDF con pipeline ibrida:
    1. Estrae immagini bitmap embedded
    2. Renderizza pagine e detecta regioni visive (figure/diagrammi) con contorni
    """
    import cv2
    import numpy as np
    
    # Usa la cartella da config se non specificata
    if output_folder is None:
        output_folder = get_images_folder()
    
    os.makedirs(output_folder, exist_ok=True)
    doc = fitz.open(pdf_path)
    images = []
    filename = os.path.splitext(os.path.basename(pdf_path))[0]
    
    for page_num in range(len(doc)):
        page = doc[page_num]
        
        # ===== STEP 1: Estrai immagini bitmap embedded =====
        image_list = page.get_images()
        bitmap_count = 0
        
        for img_idx, img in enumerate(image_list):
            xref = img[0]
            try:
                base_image = doc.extract_image(xref)
                image_bytes = base_image["image"]
                image_ext = base_image["ext"]
                
                pil_img = Image.open(BytesIO(image_bytes))
                width, height = pil_img.size
                
                if width < min_size or height < min_size:
                    continue
                
                img_filename = f"{filename}_p{page_num+1}_bitmap{img_idx+1}.{image_ext}"
                img_path = os.path.join(output_folder, img_filename)
                
                with open(img_path, "wb") as f:
                    f.write(image_bytes)
                
                images.append({
                    'page': page_num + 1,
                    'image_path': img_path,
                    'size': (width, height),
                    'source_doc': os.path.basename(pdf_path),
                    'type': 'bitmap',
                    'bbox': None
                })
                bitmap_count += 1
                
            except Exception as e:
                print(f"[WARN] Errore estrazione bitmap p{page_num+1} idx{img_idx}: {e}")
        
        # ===== STEP 2: Layout Detection per figure vettoriali =====
        if use_layout_detection:
            try:
                mat = fitz.Matrix(300/72, 300/72)
                pix = page.get_pixmap(matrix=mat)
                
                img_data = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)
                
                if pix.n == 4:
                    img_bgr = cv2.cvtColor(img_data, cv2.COLOR_RGBA2BGR)
                else:
                    img_bgr = cv2.cvtColor(img_data, cv2.COLOR_RGB2BGR)
                
                gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
                edges = cv2.Canny(gray, 50, 150)
                kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (5, 5))
                dilated = cv2.dilate(edges, kernel, iterations=2)
                
                contours, _ = cv2.findContours(dilated, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                
                region_idx = 0
                for contour in contours:
                    x, y, w, h = cv2.boundingRect(contour)
                    
                    if w < min_size or h < min_size:
                        continue
                    if w > pix.width * 0.95 or h > pix.height * 0.95:
                        continue
                    if w / h > 10 or h / w > 10:
                        continue
                    
                    padding = 10
                    x1 = max(0, x - padding)
                    y1 = max(0, y - padding)
                    x2 = min(pix.width, x + w + padding)
                    y2 = min(pix.height, y + h + padding)
                    
                    crop = img_bgr[y1:y2, x1:x2]
                    
                    img_filename = f"{filename}_p{page_num+1}_region{region_idx+1}.png"
                    img_path = os.path.join(output_folder, img_filename)
                    cv2.imwrite(img_path, crop)
                    
                    images.append({
                        'page': page_num + 1,
                        'image_path': img_path,
                        'size': (x2-x1, y2-y1),
                        'source_doc': os.path.basename(pdf_path),
                        'type': 'region_detected',
                        'bbox': (x1, y1, x2, y2)
                    })
                    region_idx += 1
                
                if region_idx > 0:
                    print(f"      ✓ Pagina {page_num+1}: {bitmap_count} bitmap + {region_idx} regioni detectate")
                
            except Exception as e:
                print(f"[WARN] Errore layout detection p{page_num+1}: {e}")
    
    doc.close()
    return images


# ===== FUNZIONI CARICAMENTO DOCUMENTI =====

def load_documents(folder):
    """Carica documenti (txt, md, pdf, docx, xlsx, xls)"""
    docs = []
    for path in sorted(glob.glob(os.path.join(folder, '*'))):
        name = os.path.basename(path)
        ext = os.path.splitext(path)[1].lower()
        try:
            if ext == '.txt' or ext == '.md':
                with open(path, 'r', encoding='utf-8') as f:
                    text = f.read()
            elif ext == '.pdf':
                reader = PdfReader(path)
                pages = [p.extract_text() or '' for p in reader.pages]
                text = "\n".join(pages)
            elif ext == '.docx':
                from docx import Document
                doc = Document(path)
                paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
                text = "\n".join(paragraphs)
            elif ext == '.xlsx':
                from openpyxl import load_workbook
                wb = load_workbook(path, data_only=True)
                all_sheets = []
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    rows = []
                    for row in ws.iter_rows(values_only=True):
                        row_text = [str(cell) for cell in row if cell is not None]
                        if row_text:
                            rows.append(" | ".join(row_text))
                    if rows:
                        all_sheets.append(f"[{sheet}]\n" + "\n".join(rows))
                text = "\n\n".join(all_sheets)
            elif ext == '.xls':
                import xlrd
                wb = xlrd.open_workbook(path)
                all_sheets = []
                for sheet in wb.sheets():
                    rows = []
                    for r in range(sheet.nrows):
                        row = [str(sheet.cell_value(r, c)) for c in range(sheet.ncols) if sheet.cell_value(r, c)]
                        if row:
                            rows.append(" | ".join(row))
                    if rows:
                        all_sheets.append(f"[{sheet.name}]\n" + "\n".join(rows))
                text = "\n\n".join(all_sheets)
            else:
                continue
            docs.append({'path': name, 'text': text.strip() if text else '', 'full_path': path})
        except Exception as e:
            print(f"[WARN] Errore leggendo {path}: {e}")
    return docs


def chunk_text(text, chunk_size=800, overlap=200):
    """Divide testo in chunk con overlap"""
    sentences = re.split(r'(?<=[\.\?\!])\s+', text)
    chunks = []
    cur = ''
    for s in sentences:
        if not cur:
            cur = s
        elif len(cur) + 1 + len(s) <= chunk_size:
            cur = cur + ' ' + s
        else:
            chunks.append(cur.strip())
            if overlap > 0:
                seed = cur[-overlap:]
                cur = (seed + ' ' + s).strip()
            else:
                cur = s
    if cur:
        chunks.append(cur.strip())
    return chunks


# ===== CONTEXTUAL RETRIEVAL =====

def contextualize_text_chunks_batch(doc_text, chunks, doc_name, batch_size=10, domain_prompt=None, client=None):
    """
    Contextual Retrieval per chunk di TESTO
    Genera contesto LLM per ogni chunk testuale
    """
    if client is not None:
        import warnings
        warnings.warn("Il parametro 'client' è deprecato. Viene usata la configurazione da config.py.", DeprecationWarning)

    if domain_prompt is None:
        domain_prompt = """You are an expert technical document analyst.
For each text chunk, provide a succinct context (1-2 sentences) that:
- Identifies the document section/topic
- Notes key technical details not explicit in the chunk
- Helps someone searching for this information to find it

Focus on technical accuracy and searchability."""

    contextualized = []
    
    for i in range(0, len(chunks), batch_size):
        batch = chunks[i:i+batch_size]
        
        chunks_xml = ""
        for idx, chunk in enumerate(batch):
            chunk_id = i + idx
            chunks_xml += f'<chunk id="{chunk_id}">\n{chunk}\n</chunk>\n\n'
        
        prompt = f"""<document name="{doc_name}">
{doc_text[:15000]}
</document>

{domain_prompt}

Chunks to contextualize:

{chunks_xml}

Return ONLY valid JSON with format:
{{
  "0": "context for chunk 0",
  "1": "context for chunk 1",
  ...
}}"""

        try:
            response_text, _ = call_llm_text(
                MODEL_PROMPT,
                system_prompt="You are a precise analyst. Return only valid JSON.",
                user_prompt=prompt,
            )
            content = response_text.strip()
            
            if content.startswith('```'):
                content = re.sub(r'^```(?:json)?\n', '', content)
                content = re.sub(r'\n```$', '', content)
            
            contexts = json.loads(content)
            
            for idx, chunk in enumerate(batch):
                chunk_id = str(i + idx)
                context = contexts.get(chunk_id, contexts.get(str(idx), ""))
                
                if context:
                    contextualized_chunk = f"CONTEXT: {context}\n\nCONTENT: {chunk}"
                else:
                    contextualized_chunk = f"CONTENT: {chunk}"
                
                contextualized.append(contextualized_chunk)
            
            print(f"  ✓ Contestualizzati chunk testo {i}-{i+len(batch)}")
            
        except Exception as e:
            print(f"  [WARN] Errore contestualizzazione batch {i}: {e}")
            for chunk in batch:
                contextualized.append(f"CONTENT: {chunk}")
    
    return contextualized


def contextualize_image_descriptions(image_descriptions, doc_name):
    """
    Contextual Retrieval per IMMAGINI
    Arricchisce le descrizioni delle immagini con contesto posizionale
    """
    contextualized = []
    
    for img_data in image_descriptions:
        description = img_data['description']
        page = img_data['page']
        img_path = img_data['image_path']
        
        context = f"Immagine tecnica dalla pagina {page} del documento '{doc_name}', salvata come {os.path.basename(img_path)}. Analizzata con Azure AI Vision."
        
        contextualized_desc = f"CONTEXT: {context}\n\nCONTENT: {description}"
        
        contextualized.append({
            **img_data,
            'description_contextualized': contextualized_desc
        })
    
    return contextualized


# ===== BUILD INDEX =====

def build_index_with_azure_vision(
    docs, 
    embed_model_name='all-MiniLM-L6-v2',
    index_path='docs_index_multimodal_contextual.faiss',
    meta_path='metadata_multimodal_contextual.json',
    chunk_size=800, 
    overlap=200,
    extract_images=True,
    analyze_images=True,
    use_text_contextualization=True,
    batch_size=10,
    domain_prompt=None,
    min_image_size=100
):
    """
    Build index con:
    - Contextual Retrieval per chunk di testo
    - Analisi immagini con Azure AI Vision
    - Contestualizzazione delle descrizioni immagini
    """
    
    # Verifica configurazione Azure Vision
    if AZURE_VISION['api_key'] == '<YOUR-API-KEY>':
        print("❌ ERRORE: Configura AZURE_VISION in config.py con endpoint e api_key validi!")
        exit(1)
    
    model = SentenceTransformer(embed_model_name)
    # Il client LLM è gestito da llm_client.py in base al provider configurato

    all_chunks = []
    metadata = []
    
    print("\n" + "="*70)
    print("RAG MULTIMODALE + AZURE AI VISION")
    print("="*70)
    print(f"🔵 Azure Vision Endpoint: {AZURE_VISION['endpoint']}")
    print(f"🔵 Features: {', '.join(AZURE_VISION.get('features', ['caption', 'tags', 'read']))}")
    
    # ===== FASE 1: PROCESSING DOCUMENTI =====
    print(f"\n[1/5] Processing {len(docs)} documenti...")
    
    for doc in tqdm(docs, desc="Documenti"):
        doc_name = doc['path']
        doc_text = doc['text']
        doc_full_path = doc['full_path']
        
        # --- TESTO: Chunking ---
        text_chunks = chunk_text(doc_text, chunk_size=chunk_size, overlap=overlap)
        print(f"\n  📄 {doc_name}: {len(text_chunks)} chunk di testo")
        
        # --- TESTO: Contextual Retrieval ---
        if use_text_contextualization and text_chunks:
            print(f"     Contestualizzazione chunk testo con LLM...")
            text_chunks_contextualized = contextualize_text_chunks_batch(
                doc_text, text_chunks, doc_name,
                batch_size=batch_size,
                domain_prompt=domain_prompt,
            )
        else:
            text_chunks_contextualized = [f"CONTENT: {c}" for c in text_chunks]
        
        # Aggiungi chunk di testo
        for i, (original, contextualized) in enumerate(zip(text_chunks, text_chunks_contextualized)):
            all_chunks.append({
                'type': 'text',
                'source': doc_name,
                'chunk_id': len(all_chunks),
                'text_original': original,
                'text_for_embedding': contextualized,
                'page': None
            })
        
        # --- IMMAGINI: Estrazione ---
        if extract_images and doc_full_path.endswith('.pdf'):
            print(f"     Estrazione immagini...")
            images = extract_images_from_pdf(doc_full_path, output_folder=get_images_folder(), min_size=min_image_size)
            print(f"     🖼️  {len(images)} immagini estratte")
            
            # Estrai il testo di ogni pagina per contesto
            page_texts = {}
            if analyze_images and images:
                print(f"     Estrazione testo pagine per contesto...")
                try:
                    reader = PdfReader(doc_full_path)
                    for page_idx, page in enumerate(reader.pages):
                        page_texts[page_idx + 1] = page.extract_text() or ''
                except Exception as e:
                    print(f"     [WARN] Errore estrazione testo pagine: {e}")
            
            # --- IMMAGINI: Analisi con Azure Vision ---
            if analyze_images and images:
                print(f"     🔵 Analisi immagini con Azure AI Vision...")
                image_descriptions = []
                
                for img in tqdm(images, desc="     - Azure Vision", leave=False):
                    page_text = page_texts.get(img['page'], '')
                    description = analyze_image_with_azure_vision(
                        img['image_path'],
                        img['page'], doc_name, page_text
                    )
                    image_descriptions.append({
                        'page': img['page'],
                        'image_path': img['image_path'],
                        'description': description,
                        'size': img['size']
                    })
                
                # --- IMMAGINI: Contestualizzazione ---
                image_descriptions_contextualized = contextualize_image_descriptions(
                    image_descriptions, doc_name
                )
                
                # Aggiungi chunk di immagini
                for img_data in image_descriptions_contextualized:
                    all_chunks.append({
                        'type': 'image',
                        'source': doc_name,
                        'chunk_id': len(all_chunks),
                        'text_original': img_data['description'],
                        'text_for_embedding': img_data['description_contextualized'],
                        'page': img_data['page'],
                        'image_path': img_data['image_path'],
                        'image_size': img_data['size']
                    })
    
    if not all_chunks:
        raise ValueError("❌ Nessun chunk generato!")
    
    # ===== FASE 2: STATISTICHE =====
    print(f"\n[2/5] Statistiche chunk...")
    text_chunks_count = sum(1 for c in all_chunks if c['type'] == 'text')
    image_chunks_count = sum(1 for c in all_chunks if c['type'] == 'image')
    print(f"  📄 Chunk testo: {text_chunks_count}")
    print(f"  🖼️  Chunk immagini: {image_chunks_count}")
    print(f"  📊 Totale: {len(all_chunks)}")
    
    # ===== FASE 3: EMBEDDINGS =====
    print(f"\n[3/5] Calcolo embeddings...")
    texts_for_embedding = [c['text_for_embedding'] for c in all_chunks]
    embeddings = model.encode(texts_for_embedding, show_progress_bar=True, convert_to_numpy=True)
    
    norms = np.linalg.norm(embeddings, axis=1, keepdims=True) + 1e-10
    embeddings = embeddings / norms
    
    # ===== FASE 4: INDEX FAISS =====
    print(f"\n[4/5] Costruzione index FAISS...")
    d = embeddings.shape[1]
    index = faiss.IndexFlatIP(d)
    index.add(embeddings.astype('float32'))
    faiss.write_index(index, index_path)
    
    # ===== FASE 5: SALVATAGGIO METADATA =====
    print(f"\n[5/5] Salvataggio metadata e chunk...")
    
    for chunk in all_chunks:
        meta = {
            'type': chunk['type'],
            'source': chunk['source'],
            'chunk_id': chunk['chunk_id'],
            'text_preview': chunk['text_original'][:200]
        }
        if chunk['type'] == 'image':
            meta['page'] = chunk['page']
            meta['image_path'] = chunk['image_path']
        metadata.append(meta)
    
    with open(meta_path, 'w', encoding='utf-8') as f:
        json.dump({'metadata': metadata}, f, ensure_ascii=False, indent=2)
    
    chunks_path = get_chunks_path()
    with open(chunks_path, 'w', encoding='utf-8') as f:
        for chunk in all_chunks:
            json.dump({
                'chunk_id': chunk['chunk_id'],
                'type': chunk['type'],
                'source': chunk['source'],
                'text_original': chunk['text_original'],
                'text_contextualized': chunk['text_for_embedding'],
                'page': chunk.get('page'),
                'image_path': chunk.get('image_path')
            }, f, ensure_ascii=False)
            f.write('\n')
    
    print("\n" + "="*70)
    print("✅ COMPLETATO!")
    print("="*70)
    print(f"📁 Index FAISS: {index_path}")
    print(f"📁 Metadata: {meta_path}")
    print(f"📁 Chunks: {chunks_path}")
    print(f"📊 Totale chunk: {len(all_chunks)}")
    print(f"   - 📄 Testo (contextual): {text_chunks_count}")
    print(f"   - 🖼️  Immagini (Azure Vision): {image_chunks_count}")
    print("="*70 + "\n")
    
    return index_path, meta_path


# ===== MAIN =====

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='RAG Multimodale + Azure AI Vision')
    parser.add_argument('--docs', type=str, default='docs', help='Cartella documenti')
    parser.add_argument('--index', type=str, default=None, help='Path index FAISS (default: da config.py)')
    parser.add_argument('--meta', type=str, default=None, help='Path metadata JSON (default: da config.py)')
    parser.add_argument('--embed_model', type=str, default=EMBEDDING_MODEL)
    parser.add_argument('--chunk_size', type=int, default=CHUNK_SIZE)
    parser.add_argument('--overlap', type=int, default=CHUNK_OVERLAP)
    parser.add_argument('--no-images', action='store_true', help='Disabilita estrazione immagini')
    parser.add_argument('--no-analyze', action='store_true', help='Estrai ma non analizzare immagini')
    parser.add_argument('--no-text-context', action='store_true', help='Disabilita contextual retrieval per testo')
    parser.add_argument('--batch_size', type=int, default=BATCH_SIZE, help='Chunk per chiamata LLM')
    parser.add_argument('--domain_prompt', type=str, help='Path file prompt personalizzato')
    parser.add_argument('--min_image_size', type=int, default=MIN_IMAGE_SIZE, help='Dimensione minima immagini (px)')
    
    args = parser.parse_args()
    
    # Crea directory output
    ensure_output_dir()
    
    # Usa paths da config se non specificati
    index_path = args.index if args.index else get_index_path()
    meta_path = args.meta if args.meta else get_metadata_path()
    
    # Carica domain prompt
    domain_prompt = None
    if args.domain_prompt and os.path.exists(args.domain_prompt):
        with open(args.domain_prompt, 'r', encoding='utf-8') as f:
            domain_prompt = f.read()
        print(f"✓ Domain prompt caricato: {args.domain_prompt}")
    
    docs = load_documents(args.docs)
    if not docs:
        print(f"❌ Nessun documento in {args.docs}")
        exit(1)
    
    build_index_with_azure_vision(
        docs,
        embed_model_name=args.embed_model,
        index_path=index_path,
        meta_path=meta_path,
        chunk_size=args.chunk_size,
        overlap=args.overlap,
        extract_images=not args.no_images,
        analyze_images=not args.no_analyze,
        use_text_contextualization=not args.no_text_context,
        batch_size=args.batch_size,
        domain_prompt=domain_prompt,
        min_image_size=args.min_image_size
    )
