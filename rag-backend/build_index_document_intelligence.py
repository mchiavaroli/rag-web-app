"""
Document Processor — Estrazione testo + immagini per LLM Wiki.

Pipeline:
1. Carica documenti (PDF via Azure Document Intelligence, DOCX, TXT, XLSX)
2. Chunk testo + Contextual Retrieval (LLM genera contesto per ogni chunk)
3. Estrazione + analisi immagini con LLM vision
4. Salva tutto in output/chunks.jsonl (letto da wiki_ingest)
"""
import os
import glob
import json
import re
import base64
from io import BytesIO
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm import tqdm
from PIL import Image
import fitz  # PyMuPDF — fallback estrazione immagini

from config import (
    MODEL_PROMPT, MODEL_IMAGE_ANALYSE,
    CHUNK_SIZE, CHUNK_OVERLAP, BATCH_SIZE, LLM_CONCURRENCY, MIN_IMAGE_SIZE,
    get_chunks_path, get_images_folder, ensure_output_dir,
)
from llm_client import call_llm_text, call_llm_with_image
from document_intelligence_extractor import DocumentIntelligenceExtractor


# ============================================================
# UTILITIES
# ============================================================

def encode_image_to_base64(image_path: str) -> str:
    with open(image_path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")


def analyze_image_with_context(image_path: str, page_num: int, source_doc: str,
                                page_text: str = None) -> str:
    """Analizza immagine con il modello configurato in MODEL_IMAGE_ANALYSE."""
    b64 = encode_image_to_base64(image_path)
    context_section = ""
    if page_text and page_text.strip():
        preview = page_text[:1500] + ("..." if len(page_text) > 1500 else "")
        context_section = f"\n\nPAGE {page_num} CONTEXT:\n{preview}\n"

    prompt = (
        f'You are an expert engineer analysing technical drawings, schematics, diagrams and charts.\n\n'
        f'Analyse this image from page {page_num} of the document "{source_doc}".{context_section}\n\n'
        "Focus on:\n"
        "1. Content type (schematic, flow diagram, chart, technical photo, etc.)\n"
        "2. Main components and identifiers (codes, abbreviations, names)\n"
        "3. Numerical values, measurements, specifications\n"
        "4. Connections, relationships, flows\n"
        "5. Text present in the image (labels, legends, annotations)\n"
        "6. Technical context and purpose\n\n"
        "Provide a complete, searchable description."
    )
    ext = os.path.splitext(image_path)[1].lower().lstrip(".")
    media_type = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
                  "gif": "image/gif", "webp": "image/webp"}.get(ext, "image/jpeg")
    try:
        return call_llm_with_image(MODEL_IMAGE_ANALYSE, b64, media_type, prompt)
    except Exception as e:
        print(f"[WARN] Analisi immagine {image_path}: {e}")
        return f"[Immagine dalla pagina {page_num}] Analisi non disponibile."


def chunk_text(text: str, chunk_size: int = 800, overlap: int = 200) -> list:
    """Divide testo in chunk con overlap."""
    sentences = re.split(r"(?<=[\.\?\!])\s+", text)
    chunks, cur = [], ""
    for s in sentences:
        if not cur:
            cur = s
        elif len(cur) + 1 + len(s) <= chunk_size:
            cur = cur + " " + s
        else:
            chunks.append(cur.strip())
            cur = (cur[-overlap:] + " " + s).strip() if overlap > 0 else s
    if cur:
        chunks.append(cur.strip())
    return chunks


def contextualize_text_chunks_batch(doc_text: str, chunks: list, doc_name: str,
                                    batch_size: int = 10, domain_prompt: str = None,
                                    concurrency: int = None) -> list:
    """
    Contextual Retrieval: genera contesto LLM per ogni chunk di testo.
    Supporta parallelismo tramite ThreadPoolExecutor.
    """
    if domain_prompt is None:
        domain_prompt = (
            "You are an expert technical document analyst.\n"
            "For each text chunk, provide a succinct context (1-2 sentences) that identifies "
            "the section/topic and helps retrieval. Focus on technical accuracy."
        )
    if concurrency is None:
        concurrency = LLM_CONCURRENCY

    doc_text_truncated = doc_text[:15000]

    def process_batch(batch_start: int):
        batch = chunks[batch_start:batch_start + batch_size]
        chunks_xml = "".join(
            f'<chunk id="{i}">\n{c}\n</chunk>\n\n' for i, c in enumerate(batch)
        )
        user_prompt = (
            f'<document name="{doc_name}">\n{doc_text_truncated}\n</document>\n\n'
            f"{domain_prompt}\n\nChunks:\n\n{chunks_xml}\n\n"
            'Return ONLY valid JSON:\n{"0": "context 0", "1": "context 1", ...}'
        )
        try:
            response_text, _ = call_llm_text(
                MODEL_PROMPT,
                "You are a precise analyst. Return only valid JSON.",
                user_prompt,
            )
            content = response_text.strip()
            if content.startswith("```"):
                content = re.sub(r"^```(?:json)?\n", "", content)
                content = re.sub(r"\n```$", "", content)
            contexts = json.loads(content)
            result = []
            for i, chunk in enumerate(batch):
                ctx = contexts.get(str(i), "")
                result.append(
                    f"CONTEXT: {ctx}\n\nCONTENT: {chunk}" if ctx else f"CONTENT: {chunk}"
                )
            print(f"  ✓ Contextualised chunks {batch_start}-{batch_start + len(batch) - 1}")
            return batch_start, result
        except Exception as e:
            print(f"  [WARN] Contestualizzazione batch {batch_start}: {e}")
            return batch_start, [f"CONTENT: {c}" for c in batch]

    batch_starts = list(range(0, len(chunks), batch_size))
    results_map: dict = {}
    with ThreadPoolExecutor(max_workers=concurrency) as executor:
        futures = {executor.submit(process_batch, s): s for s in batch_starts}
        for future in as_completed(futures):
            start, result = future.result()
            results_map[start] = result

    out = []
    for s in batch_starts:
        out.extend(results_map[s])
    return out


def contextualize_image_descriptions(image_descriptions: list, doc_name: str) -> list:
    """Arricchisce le descrizioni immagini con contesto posizionale."""
    out = []
    for img in image_descriptions:
        ctx = (
            f"Technical image from page {img['page']} of document '{doc_name}', "
            f"saved as {os.path.basename(img['image_path'])}."
        )
        out.append({
            **img,
            "description_contextualized": f"CONTEXT: {ctx}\n\nCONTENT: {img['description']}",
        })
    return out


# ============================================================
# CARICAMENTO DOCUMENTI
# ============================================================

def load_documents_with_document_intelligence(folder: str, use_doc_intelligence: bool = True) -> list:
    """
    Carica documenti dalla cartella. PDF processati con Azure Document Intelligence.
    Supporta: PDF, TXT, MD, DOCX, XLSX, XLS.
    """
    docs = []
    di_extractor = None

    if use_doc_intelligence:
        try:
            di_extractor = DocumentIntelligenceExtractor()
            print("✅ Document Intelligence pronto")
        except ValueError as e:
            print(f"⚠️  DI non configurato: {e}. Verrà usato il metodo standard (PyPDF).")

    for path in sorted(glob.glob(os.path.join(folder, "*"))):
        name = os.path.basename(path)
        ext = os.path.splitext(path)[1].lower()
        try:
            if ext in (".txt", ".md"):
                with open(path, "r", encoding="utf-8") as f:
                    text = f.read()
                docs.append({"path": name, "text": text.strip(), "full_path": path, "di_result": None})

            elif ext == ".pdf":
                if di_extractor:
                    print(f"\n📄 Processamento {name} con Document Intelligence...")
                    di_result = di_extractor.extract_from_pdf(path)
                    text = "\n\n".join(c["content"] for c in di_result["text_chunks"])
                    docs.append({"path": name, "text": text.strip(), "full_path": path, "di_result": di_result})
                    print(
                        f"   ✓ {len(di_result['text_chunks'])} chunk, "
                        f"{len(di_result['images'])} immagini, "
                        f"{len(di_result['tables'])} tabelle"
                    )
                else:
                    from pypdf import PdfReader
                    reader = PdfReader(path)
                    text = "\n".join(p.extract_text() or "" for p in reader.pages)
                    docs.append({"path": name, "text": text.strip(), "full_path": path, "di_result": None})

            elif ext == ".docx":
                from docx import Document
                doc = Document(path)
                text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
                docs.append({"path": name, "text": text, "full_path": path, "di_result": None})

            elif ext in (".xlsx", ".xls"):
                if ext == ".xlsx":
                    from openpyxl import load_workbook
                    wb = load_workbook(path, data_only=True)
                    sheets = []
                    for sheet in wb.sheetnames:
                        ws = wb[sheet]
                        rows = [" | ".join(str(c) for c in row if c is not None)
                                for row in ws.iter_rows(values_only=True)]
                        rows = [r for r in rows if r]
                        if rows:
                            sheets.append(f"[{sheet}]\n" + "\n".join(rows))
                    text = "\n\n".join(sheets)
                else:
                    import xlrd
                    wb = xlrd.open_workbook(path)
                    sheets = []
                    for sheet in wb.sheets():
                        rows = [
                            " | ".join(str(sheet.cell_value(r, c)) for c in range(sheet.ncols)
                                       if sheet.cell_value(r, c))
                            for r in range(sheet.nrows)
                        ]
                        rows = [r for r in rows if r]
                        if rows:
                            sheets.append(f"[{sheet.name}]\n" + "\n".join(rows))
                    text = "\n\n".join(sheets)
                docs.append({"path": name, "text": text, "full_path": path, "di_result": None})

            else:
                continue

        except Exception as e:
            print(f"[WARN] Errore leggendo {path}: {e}")

    return docs


# ============================================================
# PROCESSING PRINCIPALE
# ============================================================

def build_index_with_document_intelligence(
    docs: list,
    chunk_size: int = CHUNK_SIZE,
    overlap: int = CHUNK_OVERLAP,
    analyze_images: bool = True,
    use_text_contextualization: bool = True,
    batch_size: int = BATCH_SIZE,
    domain_prompt: str = None,
    min_image_size: int = MIN_IMAGE_SIZE,
    use_document_intelligence: bool = True,
) -> None:
    """
    Estrae, chunka, contestualizza e analizza immagini dai documenti.
    Salva tutto in output/chunks.jsonl — il file che wiki_ingest legge.
    """
    ensure_output_dir()
    all_chunks: list = []

    print("\n" + "=" * 60)
    print("DOCUMENT PROCESSOR — LLM Wiki Pipeline")
    print("=" * 60)

    # ===== FASE 1: PROCESSING DOCUMENTI =====
    print(f"\n[1/3] Processing {len(docs)} documenti...")

    for doc in tqdm(docs, desc="Documenti"):
        doc_name = doc["path"]
        doc_text = doc["text"]
        doc_full_path = doc["full_path"]
        di_result = doc.get("di_result")

        # --- TESTO ---
        if use_document_intelligence and di_result:
            text_chunks_raw = []
            for di_chunk in di_result["text_chunks"]:
                role = di_chunk.get("role") or "text"
                prefix = f"[{role.upper()}] " if role != "text" else ""
                text_chunks_raw.append(prefix + di_chunk["content"])
            print(f"\n  📄 {doc_name}: {len(text_chunks_raw)} chunk strutturati (DI)")
        else:
            text_chunks_raw = chunk_text(doc_text, chunk_size=chunk_size, overlap=overlap)
            print(f"\n  📄 {doc_name}: {len(text_chunks_raw)} chunk")

        if use_text_contextualization and text_chunks_raw:
            print(f"     Contestualizzazione chunk testo...")
            text_chunks_ctx = contextualize_text_chunks_batch(
                doc_text, text_chunks_raw, doc_name,
                batch_size=batch_size, domain_prompt=domain_prompt,
            )
        else:
            text_chunks_ctx = [f"CONTENT: {c}" for c in text_chunks_raw]

        for i, (orig, ctx) in enumerate(zip(text_chunks_raw, text_chunks_ctx)):
            page = None
            if use_document_intelligence and di_result and i < len(di_result["text_chunks"]):
                page = di_result["text_chunks"][i].get("page")
            all_chunks.append({
                "type": "text",
                "source": doc_name,
                "chunk_id": len(all_chunks),
                "text_original": orig,
                "text_for_embedding": ctx,
                "page": page,
            })

        # --- TABELLE (solo con Document Intelligence) ---
        if use_document_intelligence and di_result and di_result.get("tables"):
            print(f"     📊 {len(di_result['tables'])} tabelle")
            table_texts = [
                f"[TABLE - Page {t['page']}]\n{t['text']}" for t in di_result["tables"]
            ]
            if use_text_contextualization:
                table_ctxs = contextualize_text_chunks_batch(
                    doc_text, table_texts, doc_name,
                    batch_size=batch_size, domain_prompt=domain_prompt,
                )
            else:
                table_ctxs = [f"CONTENT: {t}" for t in table_texts]
            for table, table_text, table_ctx in zip(di_result["tables"], table_texts, table_ctxs):
                all_chunks.append({
                    "type": "table",
                    "source": doc_name,
                    "chunk_id": len(all_chunks),
                    "text_original": table_text,
                    "text_for_embedding": table_ctx,
                    "page": table["page"],
                    "table_id": table.get("table_id"),
                    "table_dimensions": f"{table.get('row_count', '?')}x{table.get('column_count', '?')}",
                })

        # --- IMMAGINI ---
        if use_document_intelligence and di_result and di_result.get("images"):
            images = di_result["images"]
            print(f"     🖼️  {len(images)} immagini estratte (DI)")
        elif doc_full_path.endswith(".pdf"):
            images = _extract_images_fitz(doc_full_path, get_images_folder(), min_image_size)
            print(f"     🖼️  {len(images)} immagini estratte (PyMuPDF)")
        else:
            images = []

        if analyze_images and images:
            page_texts: dict = {}
            if use_document_intelligence and di_result:
                for p in di_result["pages"]:
                    page_texts[p["page_number"]] = p["text"]
            else:
                try:
                    from pypdf import PdfReader
                    reader = PdfReader(doc_full_path)
                    for i, page in enumerate(reader.pages):
                        page_texts[i + 1] = page.extract_text() or ""
                except Exception as e:
                    print(f"     [WARN] Testo pagine: {e}")

            image_descs = []
            for img in tqdm(images, desc="     Analisi immagini", leave=False):
                desc = analyze_image_with_context(
                    img["image_path"], img["page"], doc_name,
                    page_texts.get(img["page"], ""),
                )
                image_descs.append({
                    "page": img["page"],
                    "image_path": img["image_path"],
                    "description": desc,
                    "size": img.get("size", []),
                })

            image_descs_ctx = contextualize_image_descriptions(image_descs, doc_name)
            for img_data in image_descs_ctx:
                all_chunks.append({
                    "type": "image",
                    "source": doc_name,
                    "chunk_id": len(all_chunks),
                    "text_original": img_data["description"],
                    "text_for_embedding": img_data["description_contextualized"],
                    "page": img_data["page"],
                    "image_path": img_data["image_path"],
                })

    if not all_chunks:
        raise ValueError("❌ Nessun chunk generato!")

    # ===== FASE 2: STATISTICHE =====
    text_count = sum(1 for c in all_chunks if c["type"] == "text")
    img_count = sum(1 for c in all_chunks if c["type"] == "image")
    tbl_count = sum(1 for c in all_chunks if c["type"] == "table")
    print(
        f"\n[2/3] Chunks: {text_count} testo, {tbl_count} tabelle, "
        f"{img_count} immagini = {len(all_chunks)} totale"
    )

    # ===== FASE 3: SALVATAGGIO CHUNKS JSONL =====
    print(f"\n[3/3] Salvataggio chunks...")
    chunks_file = get_chunks_path()
    with open(chunks_file, "w", encoding="utf-8") as f:
        for chunk in all_chunks:
            data = {
                "chunk_id": chunk["chunk_id"],
                "type": chunk["type"],
                "source": chunk["source"],
                "text_original": chunk["text_original"],
                "text_contextualized": chunk["text_for_embedding"],
                "page": chunk.get("page"),
            }
            if chunk["type"] == "image":
                data["image_path"] = chunk.get("image_path")
            elif chunk["type"] == "table":
                data["table_id"] = chunk.get("table_id")
                data["table_dimensions"] = chunk.get("table_dimensions")
            json.dump(data, f, ensure_ascii=False)
            f.write("\n")

    print("\n" + "=" * 60)
    print("✅ COMPLETATO!")
    print(f"📁 Chunks: {chunks_file} ({len(all_chunks)} totale)")
    print("=" * 60 + "\n")


# ============================================================
# FALLBACK: estrazione immagini bitmap via PyMuPDF
# ============================================================

def _extract_images_fitz(pdf_path: str, output_folder: str, min_size: int = 100) -> list:
    """Estrae immagini bitmap da PDF tramite PyMuPDF (usato quando DI non è configurato)."""
    os.makedirs(output_folder, exist_ok=True)
    doc = fitz.open(pdf_path)
    images = []
    filename = os.path.splitext(os.path.basename(pdf_path))[0]

    for page_num in range(len(doc)):
        page = doc[page_num]
        for img_idx, img in enumerate(page.get_images()):
            xref = img[0]
            try:
                base_image = doc.extract_image(xref)
                img_bytes = base_image["image"]
                img_ext = base_image["ext"]
                pil_img = Image.open(BytesIO(img_bytes))
                w, h = pil_img.size
                if w < min_size or h < min_size:
                    continue
                fname = f"{filename}_p{page_num + 1}_bitmap{img_idx + 1}.{img_ext}"
                fpath = os.path.join(output_folder, fname)
                with open(fpath, "wb") as f:
                    f.write(img_bytes)
                images.append({"page": page_num + 1, "image_path": fpath, "size": (w, h)})
            except Exception as e:
                print(f"[WARN] Bitmap p{page_num + 1} idx{img_idx}: {e}")

    doc.close()
    return images
