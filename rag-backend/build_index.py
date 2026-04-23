# build_index_multimodal_contextual.py
"""
RAG Multimodale + Contextual Retrieval
Combina:
- Contextual Retrieval per chunk di testo (Anthropic 2024)
- Analisi immagini
- Contestualizzazione delle descrizioni immagini
"""
import os, glob, json, re, argparse, base64
from io import BytesIO
from tqdm import tqdm
from pypdf import PdfReader
from PIL import Image
import fitz  # PyMuPDF per estrazione immagini
from sentence_transformers import SentenceTransformer
import numpy as np
import faiss
from config import (MODEL_PROMPT, MODEL_IMAGE_ANALYSE, CHUNK_SIZE, CHUNK_OVERLAP, 
                    BATCH_SIZE, MIN_IMAGE_SIZE, USE_LAYOUT_DETECTION, EMBEDDING_MODEL,
                    get_index_path, get_metadata_path, get_chunks_path, get_images_folder,
                    ensure_output_dir)
from llm_client import call_llm_text, call_llm_with_image
from rag_logger import get_logger


# ===== FUNZIONI ESTRAZIONE IMMAGINI =====

def extract_images_from_pdf(pdf_path, output_folder=None, min_size=100, use_layout_detection=True):
    """
    Estrae immagini da PDF con pipeline ibrida:
    1. Estrae immagini bitmap embedded
    2. Renderizza pagine e detecta regioni visive (figure/diagrammi) con contorni
    
    Args:
        pdf_path: Path al PDF
        output_folder: Cartella output (default: da config.py)
        min_size: Dimensione minima immagini (px)
        use_layout_detection: Se True, usa detection layout per trovare figure vettoriali
    """
    import cv2
    import numpy as np
    
    # Usa la cartella da config se non specificata
    if output_folder is None:
        output_folder = get_images_folder()
    
    os.makedirs(output_folder, exist_ok=True)
    doc = fitz.open(pdf_path)
    images = []
    filename = os.path.splitext(os.path.basename(pdf_path))[0]
    
    for page_num in range(len(doc)):
        page = doc[page_num]
        
        # ===== STEP 1: Estrai immagini bitmap embedded =====
        image_list = page.get_images()
        bitmap_count = 0
        
        for img_idx, img in enumerate(image_list):
            xref = img[0]
            try:
                base_image = doc.extract_image(xref)
                image_bytes = base_image["image"]
                image_ext = base_image["ext"]
                
                pil_img = Image.open(BytesIO(image_bytes))
                width, height = pil_img.size
                
                if width < min_size or height < min_size:
                    continue
                
                img_filename = f"{filename}_p{page_num+1}_bitmap{img_idx+1}.{image_ext}"
                img_path = os.path.join(output_folder, img_filename)
                
                with open(img_path, "wb") as f:
                    f.write(image_bytes)
                
                images.append({
                    'page': page_num + 1,
                    'image_path': img_path,
                    'size': (width, height),
                    'source_doc': os.path.basename(pdf_path),
                    'type': 'bitmap',
                    'bbox': None
                })
                bitmap_count += 1
                
            except Exception as e:
                print(f"[WARN] Errore estrazione bitmap p{page_num+1} idx{img_idx}: {e}")
        
        # ===== STEP 2: Layout Detection per figure vettoriali =====
        if use_layout_detection:
            try:
                # Render pagina a 300 DPI
                mat = fitz.Matrix(300/72, 300/72)
                pix = page.get_pixmap(matrix=mat)
                
                # Converti in array numpy per OpenCV
                img_data = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)
                
                # Converti RGB → BGR per OpenCV
                if pix.n == 4:  # RGBA
                    img_bgr = cv2.cvtColor(img_data, cv2.COLOR_RGBA2BGR)
                else:  # RGB
                    img_bgr = cv2.cvtColor(img_data, cv2.COLOR_RGB2BGR)
                
                # Grayscale per detection
                gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
                
                # Edge detection + morphology
                edges = cv2.Canny(gray, 50, 150)
                kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (5, 5))
                dilated = cv2.dilate(edges, kernel, iterations=2)
                
                # Trova contorni
                contours, _ = cv2.findContours(dilated, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                
                # Filtra e salva regioni
                region_idx = 0
                for contour in contours:
                    x, y, w, h = cv2.boundingRect(contour)
                    
                    # Filtri: dimensione minima, aspect ratio ragionevole
                    if w < min_size or h < min_size:
                        continue
                    if w > pix.width * 0.95 or h > pix.height * 0.95:  # Skip full-page boxes
                        continue
                    if w / h > 10 or h / w > 10:  # Skip troppo stretti
                        continue
                    
                    # Crop regione con padding
                    padding = 10
                    x1 = max(0, x - padding)
                    y1 = max(0, y - padding)
                    x2 = min(pix.width, x + w + padding)
                    y2 = min(pix.height, y + h + padding)
                    
                    crop = img_bgr[y1:y2, x1:x2]
                    
                    # Salva crop
                    img_filename = f"{filename}_p{page_num+1}_region{region_idx+1}.png"
                    img_path = os.path.join(output_folder, img_filename)
                    cv2.imwrite(img_path, crop)
                    
                    images.append({
                        'page': page_num + 1,
                        'image_path': img_path,
                        'size': (x2-x1, y2-y1),
                        'source_doc': os.path.basename(pdf_path),
                        'type': 'region_detected',
                        'bbox': (x1, y1, x2, y2)
                    })
                    region_idx += 1
                
                if region_idx > 0:
                    print(f"      ✓ Pagina {page_num+1}: {bitmap_count} bitmap + {region_idx} regioni detectate")
                
            except Exception as e:
                print(f"[WARN] Errore layout detection p{page_num+1}: {e}")
    
    doc.close()
    return images


def encode_image_to_base64(image_path):
    """Converte immagine in base64"""
    with open(image_path, "rb") as img_file:
        return base64.b64encode(img_file.read()).decode('utf-8')


def analyze_image_with_context(image_path, page_num, source_doc, page_text=None):
    """Analizza immagine con il modello configurato in MODEL_IMAGE_ANALYSE,
    utilizzando il testo della pagina come contesto"""

    base64_image = encode_image_to_base64(image_path)

    # Costruisci il prompt con contesto della pagina
    context_section = ""
    if page_text and page_text.strip():
        page_text_preview = page_text[:1500]
        if len(page_text) > 1500:
            page_text_preview += "..."
        context_section = f"""

📄 CONTESTO DELLA PAGINA {page_num}:
{page_text_preview}

⚠️ Usa questo contesto per identificare l'argomento specifico dell'immagine. 
Se il testo menziona argomenti, monumenti, codici o concetti tecnici, 
FOCALIZZATI su quelli nell'analisi dell'immagine.
"""

    prompt = f"""Sei un esperto ingegnere che analizza disegni tecnici, schemi, diagrammi e grafici.

Analizza questa immagine dalla pagina {page_num} del documento "{source_doc}".{context_section}

Focalizzati su:
1. Tipo di contenuto (schema elettrico, diagramma di flusso, grafico, planimetria, tabella, foto tecnica, etc.)
2. **ARGOMENTO SPECIFICO**: identifica l'argomento principale basandoti sul contesto della pagina
3. Componenti principali e loro identificatori (codici, sigle, nomi)
4. Valori numerici, misure, specifiche tecniche visibili
5. Connessioni, relazioni, flussi tra elementi
6. Testo presente nell'immagine (etichette, legende, annotazioni)
7. Contesto tecnico e scopo del disegno

Se il contesto della pagina menziona argomenti specifici, METTILI IN EVIDENZA nella descrizione.
Fornisci una descrizione completa e ricercabile che permetta di recuperare questa immagine con query specifiche."""

    ext = os.path.splitext(image_path)[1].lower()
    media_type_map = {
        '.jpg':  'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.png':  'image/png',
        '.gif':  'image/gif',
        '.webp': 'image/webp',
    }
    media_type = media_type_map.get(ext, 'image/jpeg')

    try:
        return call_llm_with_image(MODEL_IMAGE_ANALYSE, base64_image, media_type, prompt)
    except Exception as e:
        print(f"[WARN] Errore analisi per {image_path}: {e}")
        return f"[Immagine dalla pagina {page_num}] Analisi non disponibile."


# ===== FUNZIONI CARICAMENTO DOCUMENTI =====

def load_documents(folder):
    """Carica documenti (txt, md, pdf, docx, xlsx, xls)"""
    docs = []
    for path in sorted(glob.glob(os.path.join(folder, '*'))):
        name = os.path.basename(path)
        ext = os.path.splitext(path)[1].lower()
        try:
            if ext == '.txt' or ext == '.md':
                with open(path, 'r', encoding='utf-8') as f:
                    text = f.read()
            elif ext == '.pdf':
                reader = PdfReader(path)
                pages = [p.extract_text() or '' for p in reader.pages]
                text = "\n".join(pages)
            elif ext == '.docx':
                from docx import Document
                doc = Document(path)
                paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
                text = "\n".join(paragraphs)
            elif ext == '.xlsx':
                from openpyxl import load_workbook
                wb = load_workbook(path, data_only=True)
                all_sheets = []
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    rows = []
                    for row in ws.iter_rows(values_only=True):
                        row_text = [str(cell) for cell in row if cell is not None]
                        if row_text:
                            rows.append(" | ".join(row_text))
                    if rows:
                        all_sheets.append(f"[{sheet}]\n" + "\n".join(rows))
                text = "\n\n".join(all_sheets)
            elif ext == '.xls':
                import xlrd
                wb = xlrd.open_workbook(path)
                all_sheets = []
                for sheet in wb.sheets():
                    rows = []
                    for r in range(sheet.nrows):
                        row = [str(sheet.cell_value(r, c)) for c in range(sheet.ncols) if sheet.cell_value(r, c)]
                        if row:
                            rows.append(" | ".join(row))
                    if rows:
                        all_sheets.append(f"[{sheet.name}]\n" + "\n".join(rows))
                text = "\n\n".join(all_sheets)
            else:
                continue
            # Aggiungi documento anche se il testo è vuoto (può avere immagini)
            docs.append({'path': name, 'text': text.strip() if text else '', 'full_path': path})
        except Exception as e:
            print(f"[WARN] Errore leggendo {path}: {e}")
    return docs


def chunk_text(text, chunk_size=800, overlap=200):
    """Divide testo in chunk con overlap"""
    sentences = re.split(r'(?<=[\.\?\!])\s+', text)
    chunks = []
    cur = ''
    for s in sentences:
        if not cur:
            cur = s
        elif len(cur) + 1 + len(s) <= chunk_size:
            cur = cur + ' ' + s
        else:
            chunks.append(cur.strip())
            if overlap > 0:
                seed = cur[-overlap:]
                cur = (seed + ' ' + s).strip()
            else:
                cur = s
    if cur:
        chunks.append(cur.strip())
    return chunks


# ===== CONTEXTUAL RETRIEVAL =====

def contextualize_text_chunks_batch(doc_text, chunks, doc_name, batch_size=10, domain_prompt=None):
    """
    Contextual Retrieval per chunk di TESTO.
    Genera contesto LLM per ogni chunk testuale usando il modello configurato in MODEL_IMAGE_ANALYSE.
    """
    if domain_prompt is None:
        domain_prompt = """You are an expert technical document analyst.
For each text chunk, provide a succinct context (1-2 sentences) that:
- Identifies the document section/topic
- Notes key technical details not explicit in the chunk
- Helps someone searching for this information to find it

Focus on technical accuracy and searchability."""

    contextualized = []

    for i in range(0, len(chunks), batch_size):
        batch = chunks[i:i+batch_size]

        chunks_xml = ""
        for idx, chunk in enumerate(batch):
            chunk_id = i + idx
            chunks_xml += f'<chunk id="{chunk_id}">\n{chunk}\n</chunk>\n\n'

        user_prompt = f"""<document name="{doc_name}">
{doc_text[:15000]}
</document>

{domain_prompt}

Chunks to contextualize:

{chunks_xml}

Return ONLY valid JSON with format:
{{
  "0": "context for chunk 0",
  "1": "context for chunk 1",
  ...
}}"""

        try:
            response_text, _ = call_llm_text(
                MODEL_IMAGE_ANALYSE,
                system_prompt="You are a precise analyst. Return only valid JSON.",
                user_prompt=user_prompt,
            )
            content = response_text.strip()

            if content.startswith('```'):
                content = re.sub(r'^```(?:json)?\n', '', content)
                content = re.sub(r'\n```$', '', content)

            contexts = json.loads(content)

            for idx, chunk in enumerate(batch):
                chunk_id = str(i + idx)
                context = contexts.get(chunk_id, contexts.get(str(idx), ""))
                contextualized_chunk = f"CONTEXT: {context}\n\nCONTENT: {chunk}" if context else f"CONTENT: {chunk}"
                contextualized.append(contextualized_chunk)

            print(f"  ✓ Contestualizzati chunk testo {i}-{i+len(batch)}")

        except Exception as e:
            print(f"  [WARN] Errore contestualizzazione batch {i}: {e}")
            for chunk in batch:
                contextualized.append(f"CONTENT: {chunk}")

    return contextualized


def contextualize_image_descriptions(image_descriptions, doc_name):
    """
    Contextual Retrieval per IMMAGINI.
    Arricchisce le descrizioni delle immagini con contesto posizionale.
    """
    contextualized = []

    for img_data in image_descriptions:
        description = img_data['description']
        page = img_data['page']
        img_path = img_data['image_path']

        context = f"Immagine tecnica dalla pagina {page} del documento '{doc_name}', salvata come {os.path.basename(img_path)}."
        contextualized_desc = f"CONTEXT: {context}\n\nCONTENT: {description}"

        contextualized.append({
            **img_data,
            'description_contextualized': contextualized_desc
        })

    return contextualized


# ===== BUILD INDEX =====

def build_index_multimodal_contextual(
    docs,
    embed_model_name='all-MiniLM-L6-v2',
    index_path='docs_index_multimodal_contextual.faiss',
    meta_path='metadata_multimodal_contextual.json',
    chunk_size=800,
    overlap=200,
    extract_images=True,
    analyze_images=True,
    use_text_contextualization=True,
    batch_size=10,
    domain_prompt=None,
    min_image_size=100
):
    """
    Build index con:
    - Contextual Retrieval per chunk di testo
    - Analisi immagini con il modello configurato in MODEL_IMAGE_ANALYSE
    - Contestualizzazione delle descrizioni immagini
    """

    model = SentenceTransformer(embed_model_name)

    all_chunks = []  # Lista unificata di tutti i chunk (testo + immagini)
    metadata = []

    # === LOGGING: Inizializza tracking ===
    logger = get_logger()
    log_context = logger.log_indexing_start(docs)
    llm_calls_contextualization = 0
    llm_calls_image_analysis = 0
    all_images_extracted = []

    print("\n" + "="*70)
    print("RAG MULTIMODALE + CONTEXTUAL RETRIEVAL")
    print("="*70)

    # ===== FASE 1: PROCESSING DOCUMENTI =====
    print(f"\n[1/5] Processing {len(docs)} documenti...")

    for doc in tqdm(docs, desc="Documenti"):
        doc_name = doc['path']
        doc_text = doc['text']
        doc_full_path = doc['full_path']

        # --- TESTO: Chunking ---
        text_chunks = chunk_text(doc_text, chunk_size=chunk_size, overlap=overlap)
        print(f"\n  📄 {doc_name}: {len(text_chunks)} chunk di testo")

        # --- TESTO: Contextual Retrieval ---
        if use_text_contextualization and text_chunks:
            print(f"     Contestualizzazione chunk testo...")
            text_chunks_contextualized = contextualize_text_chunks_batch(
                doc_text, text_chunks, doc_name,
                batch_size=batch_size,
                domain_prompt=domain_prompt,
            )
            # LOG: conta chiamate LLM per contestualizzazione (batch)
            llm_calls_contextualization += (len(text_chunks) + batch_size - 1) // batch_size
        else:
            text_chunks_contextualized = [f"CONTENT: {c}" for c in text_chunks]

        # Aggiungi chunk di testo
        for i, (original, contextualized) in enumerate(zip(text_chunks, text_chunks_contextualized)):
            all_chunks.append({
                'type': 'text',
                'source': doc_name,
                'chunk_id': len(all_chunks),
                'text_original': original,
                'text_for_embedding': contextualized,
                'page': None
            })

        # --- IMMAGINI: Estrazione ---
        if extract_images and doc_full_path.endswith('.pdf'):
            print(f"     Estrazione immagini...")
            images = extract_images_from_pdf(doc_full_path, output_folder=get_images_folder(), min_size=min_image_size)
            print(f"     🖼️  {len(images)} immagini estratte")

            # Estrai il testo di ogni pagina per contesto
            page_texts = {}
            if analyze_images and images:
                print(f"     Estrazione testo pagine per contesto...")
                try:
                    reader = PdfReader(doc_full_path)
                    for page_idx, page in enumerate(reader.pages):
                        page_texts[page_idx + 1] = page.extract_text() or ''
                except Exception as e:
                    print(f"     [WARN] Errore estrazione testo pagine: {e}")

            # --- IMMAGINI: Analisi ---
            if analyze_images and images:
                print(f"     Analisi immagini con contesto pagina...")
                image_descriptions = []

                for img in tqdm(images, desc="     - Analisi immagini", leave=False):
                    page_text = page_texts.get(img['page'], '')
                    description = analyze_image_with_context(
                        img['image_path'],
                        img['page'], doc_name, page_text
                    )
                    image_descriptions.append({
                        'page': img['page'],
                        'image_path': img['image_path'],
                        'description': description,
                        'size': img['size']
                    })
                    # LOG: conta chiamate LLM per analisi immagini
                    llm_calls_image_analysis += 1
                    # LOG: salva info immagine
                    all_images_extracted.append({
                        'path': img['image_path'],
                        'page': img['page'],
                        'size': list(img['size']),
                        'type': img.get('type', 'unknown')
                    })

                # --- IMMAGINI: Contestualizzazione ---
                image_descriptions_contextualized = contextualize_image_descriptions(
                    image_descriptions, doc_name
                )

                # Aggiungi chunk di immagini
                for img_data in image_descriptions_contextualized:
                    all_chunks.append({
                        'type': 'image',
                        'source': doc_name,
                        'chunk_id': len(all_chunks),
                        'text_original': img_data['description'],
                        'text_for_embedding': img_data['description_contextualized'],
                        'page': img_data['page'],
                        'image_path': img_data['image_path'],
                        'image_size': img_data['size']
                    })

    if not all_chunks:
        raise ValueError("❌ Nessun chunk generato!")

    # ===== FASE 2: STATISTICHE =====
    print(f"\n[2/5] Statistiche chunk...")
    text_chunks_count = sum(1 for c in all_chunks if c['type'] == 'text')
    image_chunks_count = sum(1 for c in all_chunks if c['type'] == 'image')
    print(f"  📄 Chunk testo: {text_chunks_count}")
    print(f"  🖼️  Chunk immagini: {image_chunks_count}")
    print(f"  📊 Totale: {len(all_chunks)}")

    # ===== FASE 3: EMBEDDINGS =====
    print(f"\n[3/5] Calcolo embeddings...")
    texts_for_embedding = [c['text_for_embedding'] for c in all_chunks]
    embeddings = model.encode(texts_for_embedding, show_progress_bar=True, convert_to_numpy=True)

    # Normalizza
    norms = np.linalg.norm(embeddings, axis=1, keepdims=True) + 1e-10
    embeddings = embeddings / norms

    # ===== FASE 4: INDEX FAISS =====
    print(f"\n[4/5] Costruzione index FAISS...")
    d = embeddings.shape[1]
    index = faiss.IndexFlatIP(d)
    index.add(embeddings.astype('float32'))
    faiss.write_index(index, index_path)

    # ===== FASE 5: SALVATAGGIO METADATA =====
    print(f"\n[5/5] Salvataggio metadata e chunk...")

    for chunk in all_chunks:
        meta = {
            'type': chunk['type'],
            'source': chunk['source'],
            'chunk_id': chunk['chunk_id'],
            'text_preview': chunk['text_original'][:200]
        }
        if chunk['type'] == 'image':
            meta['page'] = chunk['page']
            meta['image_path'] = chunk['image_path']
        metadata.append(meta)

    with open(meta_path, 'w', encoding='utf-8') as f:
        json.dump({'metadata': metadata}, f, ensure_ascii=False, indent=2)

    # Salva chunk completi
    chunks_path = get_chunks_path()
    with open(chunks_path, 'w', encoding='utf-8') as f:
        for chunk in all_chunks:
            json.dump({
                'chunk_id': chunk['chunk_id'],
                'type': chunk['type'],
                'source': chunk['source'],
                'text_original': chunk['text_original'],
                'text_contextualized': chunk['text_for_embedding'],
                'page': chunk.get('page'),
                'image_path': chunk.get('image_path')
            }, f, ensure_ascii=False)
            f.write('\n')

    print("\n" + "="*70)
    print("✅ COMPLETATO!")
    print("="*70)
    print(f"📁 Index FAISS: {index_path}")
    print(f"📁 Metadata: {meta_path}")
    print(f"📁 Chunks: {chunks_path}")
    print(f"📊 Totale chunk: {len(all_chunks)}")
    print(f"   - 📄 Testo (contextual): {text_chunks_count}")
    print(f"   - 🖼️  Immagini (vision+contextual): {image_chunks_count}")
    print("="*70 + "\n")

    # === LOGGING: Salva log indicizzazione ===
    try:
        total_pages = 0
        for doc in docs:
            if doc['full_path'].endswith('.pdf'):
                try:
                    reader = PdfReader(doc['full_path'])
                    total_pages += len(reader.pages)
                except:
                    pass

        logger.log_indexing_complete(
            context=log_context,
            document_info={
                "filename": docs[0]['path'] if len(docs) == 1 else f"{len(docs)} documenti",
                "pages": total_pages,
                "file_type": "multi" if len(docs) > 1 else os.path.splitext(docs[0]['path'])[1]
            },
            chunks_info={
                "total": len(all_chunks),
                "text": text_chunks_count,
                "images": image_chunks_count
            },
            images_info=all_images_extracted,
            llm_calls_info={
                "contextualization": llm_calls_contextualization,
                "image_analysis": llm_calls_image_analysis,
                "total_calls": llm_calls_contextualization + llm_calls_image_analysis
            },
            index_path=index_path,
            model_info={
                "text_model": MODEL_IMAGE_ANALYSE.get("deployment_name", str(MODEL_IMAGE_ANALYSE)),
                "image_model": MODEL_IMAGE_ANALYSE.get("deployment_name", str(MODEL_IMAGE_ANALYSE))
            }
        )
        print(f"📝 Log indicizzazione salvato in: output/logs/indexing_logs.jsonl")
    except Exception as e:
        print(f"[WARN] Errore salvataggio log: {e}")

    return index_path, meta_path


# ===== MAIN =====

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='RAG Multimodale + Contextual Retrieval')
    parser.add_argument('--docs', type=str, default='docs', help='Cartella documenti')
    parser.add_argument('--index', type=str, default=None, help='Path index FAISS (default: da config.py)')
    parser.add_argument('--meta', type=str, default=None, help='Path metadata JSON (default: da config.py)')
    parser.add_argument('--embed_model', type=str, default=EMBEDDING_MODEL)
    parser.add_argument('--chunk_size', type=int, default=CHUNK_SIZE)
    parser.add_argument('--overlap', type=int, default=CHUNK_OVERLAP)
    parser.add_argument('--no-images', action='store_true', help='Disabilita estrazione immagini')
    parser.add_argument('--no-analyze', action='store_true', help='Estrai ma non analizzare immagini')
    parser.add_argument('--no-text-context', action='store_true', help='Disabilita contextual retrieval per testo')
    parser.add_argument('--batch_size', type=int, default=BATCH_SIZE, help='Chunk per chiamata LLM')
    parser.add_argument('--domain_prompt', type=str, help='Path file prompt personalizzato')
    parser.add_argument('--min_image_size', type=int, default=MIN_IMAGE_SIZE, help='Dimensione minima immagini (px)')

    args = parser.parse_args()

    # Crea directory output
    ensure_output_dir()

    # Usa paths da config se non specificati
    index_path = args.index if args.index else get_index_path()
    meta_path = args.meta if args.meta else get_metadata_path()

    # Carica domain prompt
    domain_prompt = None
    if args.domain_prompt and os.path.exists(args.domain_prompt):
        with open(args.domain_prompt, 'r', encoding='utf-8') as f:
            domain_prompt = f.read()
        print(f"✓ Domain prompt caricato: {args.domain_prompt}")        

    docs = load_documents(args.docs)
    if not docs:
        print(f"❌ Nessun documento in {args.docs}")
        exit(1)

    build_index_multimodal_contextual(
        docs,
        embed_model_name=args.embed_model,
        index_path=index_path,
        meta_path=meta_path,
        chunk_size=args.chunk_size,
        overlap=args.overlap,
        extract_images=not args.no_images,
        analyze_images=not args.no_analyze,
        use_text_contextualization=not args.no_text_context,
        batch_size=args.batch_size,
        domain_prompt=domain_prompt,
        min_image_size=args.min_image_size
    )